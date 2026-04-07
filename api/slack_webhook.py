"""
Slack Events API Webhook for Vercel Serverless Functions
メンションを受け取り、Excelを更新してSlackに返信する
"""

import os
import re
import json
import datetime
import base64
import hmac
import hashlib
import tempfile
import time

import requests
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from http.server import BaseHTTPRequestHandler

# ── 設定 ──
SLACK_BOT_TOKEN = os.environ.get("SLACK_BOT_TOKEN", "")
SLACK_SIGNING_SECRET = os.environ.get("SLACK_SIGNING_SECRET", "")
GITHUB_TOKEN = os.environ.get("GITHUB_TOKEN", "")
GITHUB_REPO = os.environ.get("GITHUB_REPO", "muuuuum/ai-news")
EXCEL_FILENAME = "stakeholder_list.xlsx"

HEADERS_LIST = ["No.", "会社名", "名前", "協業", "クライアント見込み", "重要度", "メモ", "登録日"]


# ── Slack API ──

def slack_post(channel, text, thread_ts=None):
    """Slackにメッセージを投稿"""
    data = {"channel": channel, "text": text, "mrkdwn": True}
    if thread_ts:
        data["thread_ts"] = thread_ts
    resp = requests.post(
        "https://slack.com/api/chat.postMessage",
        headers={"Authorization": f"Bearer {SLACK_BOT_TOKEN}", "Content-Type": "application/json"},
        json=data,
    )
    return resp.json().get("ok", False)


def verify_slack_request(headers, body):
    """Slackリクエストの署名を検証"""
    if not SLACK_SIGNING_SECRET:
        return True  # 開発時はスキップ
    timestamp = headers.get("x-slack-request-timestamp", "")
    if abs(time.time() - int(timestamp)) > 60 * 5:
        return False
    sig_basestring = f"v0:{timestamp}:{body}"
    my_sig = "v0=" + hmac.new(
        SLACK_SIGNING_SECRET.encode(), sig_basestring.encode(), hashlib.sha256
    ).hexdigest()
    slack_sig = headers.get("x-slack-signature", "")
    return hmac.compare_digest(my_sig, slack_sig)


# ── GitHub Excel 永続化 ──

def github_api(method, path, data=None):
    url = f"https://api.github.com/repos/{GITHUB_REPO}/contents/{path}"
    headers = {
        "Authorization": f"Bearer {GITHUB_TOKEN}",
        "Accept": "application/vnd.github.v3+json",
    }
    if method == "GET":
        return requests.get(url, headers=headers)
    elif method == "PUT":
        return requests.put(url, headers=headers, json=data)


def download_excel():
    resp = github_api("GET", EXCEL_FILENAME)
    if resp.status_code == 200:
        content = base64.b64decode(resp.json()["content"])
        path = os.path.join(tempfile.gettempdir(), EXCEL_FILENAME)
        with open(path, "wb") as f:
            f.write(content)
        return path
    return None


def upload_excel(local_path):
    with open(local_path, "rb") as f:
        content = base64.b64encode(f.read()).decode()
    resp = github_api("GET", EXCEL_FILENAME)
    sha = resp.json().get("sha") if resp.status_code == 200 else None
    data = {
        "message": f"Update stakeholder list ({datetime.datetime.now().strftime('%Y-%m-%d %H:%M')})",
        "content": content,
    }
    if sha:
        data["sha"] = sha
    github_api("PUT", EXCEL_FILENAME, data)


def get_excel_path():
    path = os.path.join(tempfile.gettempdir(), EXCEL_FILENAME)
    downloaded = download_excel()
    if downloaded:
        return downloaded
    init_excel(path)
    return path


# ── Excel操作 ──

def init_excel(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ステークホルダーリスト"
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    for col, header in enumerate(HEADERS_LIST, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    widths = {"A": 6, "B": 25, "C": 15, "D": 10, "E": 18, "F": 10, "G": 40, "H": 14}
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width
    ws.auto_filter.ref = "A1:H1"
    wb.save(path)


def add_entry(company, name, collaboration, prospect, importance, memo):
    path = get_excel_path()
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    max_no = 0
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0] and isinstance(row[0], (int, float)):
            max_no = max(max_no, int(row[0]))
    no = max_no + 1
    jst = datetime.timezone(datetime.timedelta(hours=9))
    today = datetime.datetime.now(jst).strftime("%Y-%m-%d")
    new_row = [no, company, name, collaboration, prospect, importance, memo, today]
    row_num = ws.max_row + 1
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    for col, value in enumerate(new_row, 1):
        cell = ws.cell(row=row_num, column=col, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center")
    ws.auto_filter.ref = f"A1:H{row_num}"
    wb.save(path)
    upload_excel(path)
    return no


def delete_entry(company=None, name=None):
    path = get_excel_path()
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    deleted = []
    rows_to_delete = []
    for row in range(2, ws.max_row + 1):
        row_company = str(ws.cell(row=row, column=2).value or "")
        row_name = str(ws.cell(row=row, column=3).value or "")
        match = False
        if company and name:
            match = company in row_company or name in row_name
        elif company:
            match = company in row_company
        elif name:
            match = name in row_name
        if match:
            deleted.append(f"No.{ws.cell(row=row, column=1).value} {row_company} - {row_name}")
            rows_to_delete.append(row)
    for row in reversed(rows_to_delete):
        ws.delete_rows(row)
    for i, row in enumerate(range(2, ws.max_row + 1), 1):
        ws.cell(row=row, column=1, value=i)
    if rows_to_delete:
        ws.auto_filter.ref = f"A1:H{ws.max_row}"
    wb.save(path)
    upload_excel(path)
    return deleted


def get_all_entries():
    path = get_excel_path()
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    entries = []
    for row in range(2, ws.max_row + 1):
        entries.append({
            "no": ws.cell(row=row, column=1).value,
            "company": ws.cell(row=row, column=2).value or "",
            "name": ws.cell(row=row, column=3).value or "",
            "collaboration": ws.cell(row=row, column=4).value or "",
            "prospect": ws.cell(row=row, column=5).value or "",
            "importance": ws.cell(row=row, column=6).value or "",
            "memo": ws.cell(row=row, column=7).value or "",
            "date": ws.cell(row=row, column=8).value or "",
        })
    return entries


# ── メッセージ解析 ──

def parse_add(text):
    if not re.search(r"追加|登録|add", text, re.IGNORECASE):
        return None
    fields = {}
    patterns = {
        "company": r"(?:会社名|会社|企業)[：:\s]+(.+)",
        "name": r"(?:名前|氏名|担当者)[：:\s]+(.+)",
        "collaboration": r"(?:協業|協業か)[：:\s]+(.+)",
        "prospect": r"(?:クライアント見込み|見込み|クライアント)[：:\s]+(.+)",
        "importance": r"(?:重要度|優先度)[：:\s]+(.+)",
        "memo": r"(?:メモ|備考|ノート)[：:\s]+(.+)",
    }
    for key, pattern in patterns.items():
        match = re.search(pattern, text)
        if match:
            fields[key] = match.group(1).strip()
    if "company" in fields or "name" in fields:
        return fields
    return None


def parse_delete(text):
    if not re.search(r"削除|remove|delete", text, re.IGNORECASE):
        return None
    company_match = re.search(r"(?:会社名|会社)[：:\s]+(.+)", text)
    name_match = re.search(r"(?:名前|氏名)[：:\s]+(.+)", text)
    if company_match or name_match:
        return {
            "company": company_match.group(1).strip() if company_match else None,
            "name": name_match.group(1).strip() if name_match else None,
        }
    simple = re.search(r"(?:削除|remove|delete)\s+(.+)", text, re.IGNORECASE)
    if simple:
        keyword = simple.group(1).strip()
        return {"company": keyword, "name": keyword}
    return None


# ── イベント処理 ──

def process_event(event):
    """Slackイベントを処理"""
    text = event.get("text", "")
    channel = event.get("channel", "")
    thread_ts = event.get("thread_ts", event.get("ts"))

    clean = re.sub(r"<@[A-Z0-9]+>", "", text).strip()

    # 追加
    add_data = parse_add(clean)
    if add_data:
        no = add_entry(
            add_data.get("company", ""),
            add_data.get("name", ""),
            add_data.get("collaboration", ""),
            add_data.get("prospect", ""),
            add_data.get("importance", "通常"),
            add_data.get("memo", ""),
        )
        slack_post(channel, (
            f"✅ ステークホルダーを登録しました！\n\n"
            f"*No.{no}*\n"
            f"• 会社名: {add_data.get('company', '')}\n"
            f"• 名前: {add_data.get('name', '')}\n"
            f"• 協業: {add_data.get('collaboration', '')}\n"
            f"• クライアント見込み: {add_data.get('prospect', '')}\n"
            f"• 重要度: {add_data.get('importance', '通常')}\n"
            f"• メモ: {add_data.get('memo', '')}"
        ), thread_ts)
        return

    # 削除
    del_data = parse_delete(clean)
    if del_data:
        deleted = delete_entry(del_data.get("company"), del_data.get("name"))
        if deleted:
            items = "\n".join(f"• {d}" for d in deleted)
            slack_post(channel, f"🗑️ 以下を削除しました:\n{items}", thread_ts)
        else:
            keyword = del_data.get("company") or del_data.get("name") or "（不明）"
            slack_post(channel, f"⚠️ 「{keyword}」に該当するエントリが見つかりませんでした。", thread_ts)
        return

    # 一覧
    if re.search(r"一覧|リスト|list|表示|確認", clean, re.IGNORECASE):
        entries = get_all_entries()
        if not entries:
            slack_post(channel, "📋 ステークホルダーリストは空です。", thread_ts)
            return
        lines = ["📋 *ステークホルダーリスト*\n"]
        for e in entries:
            imp = e["importance"]
            icon = "🔴" if imp in ["高", "最高"] else "🟡" if imp == "通常" else "⚪"
            lines.append(
                f"{icon} *No.{e['no']}* {e['company']} / {e['name']}\n"
                f"   協業: {e['collaboration']} | 見込み: {e['prospect']} | 重要度: {imp}\n"
                f"   メモ: {e['memo']}"
            )
        lines.append(f"\n合計: {len(entries)}件")
        slack_post(channel, "\n".join(lines), thread_ts)
        return

    # ヘルプ
    slack_post(channel, (
        "📝 *営業リスト管理Bot* — 使い方\n\n"
        "*【追加】*\n```\n@営業リストBot 追加\n会社名: 株式会社ABC\n名前: 田中太郎\n"
        "協業: ○\nクライアント見込み: ◎\n重要度: 高\nメモ: 来月打ち合わせ予定```\n\n"
        "*【削除】*\n```\n@営業リストBot 削除 株式会社ABC```\n\n"
        "*【一覧表示】*\n```\n@営業リストBot 一覧```"
    ), thread_ts)


# ── Vercel Serverless Handler ──

class handler(BaseHTTPRequestHandler):
    def do_POST(self):
        content_length = int(self.headers.get("Content-Length", 0))
        body = self.rfile.read(content_length).decode("utf-8")

        try:
            payload = json.loads(body)
        except json.JSONDecodeError:
            self.send_response(400)
            self.end_headers()
            return

        # Slack URL Verification (初回設定時)
        if payload.get("type") == "url_verification":
            self.send_response(200)
            self.send_header("Content-Type", "application/json")
            self.end_headers()
            self.wfile.write(json.dumps({"challenge": payload["challenge"]}).encode())
            return

        # イベント処理
        if payload.get("type") == "event_callback":
            event = payload.get("event", {})
            # app_mention イベントのみ処理
            if event.get("type") == "app_mention":
                # bot自身の投稿は無視
                if event.get("bot_id"):
                    self.send_response(200)
                    self.end_headers()
                    return
                process_event(event)

        self.send_response(200)
        self.end_headers()
