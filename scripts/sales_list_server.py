#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
営業リストBot - Socket Mode（即時応答）
Slack Socket Mode で常時接続し、メンションに即座に返信する。
ExcelファイルはGitHubリポジトリに保存し、永続化する。
"""

import os
import re
import json
import datetime
import base64
import tempfile

import requests
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from slack_bolt import App
from slack_bolt.adapter.socket_mode import SocketModeHandler

# ── 設定 ──
SLACK_BOT_TOKEN = os.environ.get("SLACK_BOT_TOKEN", "")
SLACK_APP_TOKEN = os.environ.get("SLACK_APP_TOKEN", "")
GITHUB_TOKEN = os.environ.get("GITHUB_TOKEN", "")
GITHUB_REPO = os.environ.get("GITHUB_REPO", "muuuuum/ai-news")
EXCEL_FILENAME = "stakeholder_list.xlsx"
CHANNEL_ID = "C0AR48QHRGB"

HEADERS_LIST = ["No.", "会社名", "名前", "協業", "クライアント見込み", "重要度", "メモ", "登録日"]

app = App(token=SLACK_BOT_TOKEN)


# ── GitHub Excel 永続化 ──

def github_api(method, path, data=None):
    """GitHub API を呼び出す"""
    url = f"https://api.github.com/repos/{GITHUB_REPO}/contents/{path}"
    headers = {
        "Authorization": f"Bearer {GITHUB_TOKEN}",
        "Accept": "application/vnd.github.v3+json",
    }
    if method == "GET":
        resp = requests.get(url, headers=headers)
    elif method == "PUT":
        resp = requests.put(url, headers=headers, json=data)
    return resp


def download_excel():
    """GitHubからExcelをダウンロード"""
    resp = github_api("GET", EXCEL_FILENAME)
    if resp.status_code == 200:
        content = base64.b64decode(resp.json()["content"])
        path = os.path.join(tempfile.gettempdir(), EXCEL_FILENAME)
        with open(path, "wb") as f:
            f.write(content)
        print(f"✅ Excel downloaded from GitHub")
        return path
    return None


def upload_excel(local_path):
    """ExcelをGitHubにアップロード"""
    with open(local_path, "rb") as f:
        content = base64.b64encode(f.read()).decode()

    # 既存ファイルのSHAを取得
    resp = github_api("GET", EXCEL_FILENAME)
    sha = resp.json().get("sha") if resp.status_code == 200 else None

    data = {
        "message": f"Update stakeholder list ({datetime.datetime.now().strftime('%Y-%m-%d %H:%M')})",
        "content": content,
    }
    if sha:
        data["sha"] = sha

    resp = github_api("PUT", EXCEL_FILENAME, data)
    if resp.status_code in (200, 201):
        print(f"✅ Excel uploaded to GitHub")
    else:
        print(f"❌ Upload failed: {resp.status_code} {resp.text[:200]}")


def get_excel_path():
    """Excelのローカルパスを取得（なければGitHubからDL・なければ新規作成）"""
    path = os.path.join(tempfile.gettempdir(), EXCEL_FILENAME)
    if not os.path.exists(path):
        downloaded = download_excel()
        if downloaded:
            return downloaded
        init_excel(path)
    return path


# ── Excel操作 ──

def init_excel(path):
    """Excelファイルを初期化"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ステークホルダーリスト"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col, header in enumerate(HEADERS_LIST, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    widths = {"A": 6, "B": 25, "C": 15, "D": 10, "E": 18, "F": 10, "G": 40, "H": 14}
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    ws.auto_filter.ref = "A1:H1"
    wb.save(path)
    print(f"✅ Excel created: {path}")


def add_entry(company, name, collaboration, prospect, importance, memo):
    """追加"""
    path = get_excel_path()
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    max_no = 0
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0] and isinstance(row[0], (int, float)):
            max_no = max(max_no, int(row[0]))
    no = max_no + 1

    jst = datetime.timezone(datetime.timedelta(hours=9))
    today = datetime.datetime.now(jst).strftime("%Y-%m-%d")
    new_row = [no, company, name, collaboration, prospect, importance, memo, today]
    row_num = ws.max_row + 1

    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    for col, value in enumerate(new_row, 1):
        cell = ws.cell(row=row_num, column=col, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center")

    ws.auto_filter.ref = f"A1:H{row_num}"
    wb.save(path)
    upload_excel(path)
    return no


def delete_entry(company=None, name=None):
    """削除"""
    path = get_excel_path()
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    deleted = []
    rows_to_delete = []

    for row in range(2, ws.max_row + 1):
        row_company = str(ws.cell(row=row, column=2).value or "")
        row_name = str(ws.cell(row=row, column=3).value or "")

        match = False
        if company and name:
            match = company in row_company or name in row_name
        elif company:
            match = company in row_company
        elif name:
            match = row_name

        if match:
            deleted.append(f"No.{ws.cell(row=row, column=1).value} {row_company} - {row_name}")
            rows_to_delete.append(row)

    for row in reversed(rows_to_delete):
        ws.delete_rows(row)

    for i, row in enumerate(range(2, ws.max_row + 1), 1):
        ws.cell(row=row, column=1, value=i)

    if rows_to_delete:
        ws.auto_filter.ref = f"A1:H{ws.max_row}"

    wb.save(path)
    upload_excel(path)
    return deleted


def get_all_entries():
    """全件取得"""
    path = get_excel_path()
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    entries = []
    for row in range(2, ws.max_row + 1):
        entries.append({
            "no": ws.cell(row=row, column=1).value,
            "company": ws.cell(row=row, column=2).value or "",
            "name": ws.cell(row=row, column=3).value or "",
            "collaboration": ws.cell(row=row, column=4).value or "",
            "prospect": ws.cell(row=row, column=5).value or "",
            "importance": ws.cell(row=row, column=6).value or "",
            "memo": ws.cell(row=row, column=7).value or "",
            "date": ws.cell(row=row, column=8).value or "",
        })
    return entries


# ── メッセージ解析 ──

def parse_add(text):
    """追加コマンド解析"""
    if not re.search(r"追加|登録|add", text, re.IGNORECASE):
        return None
    fields = {}
    patterns = {
        "company": r"(?:会社名|会社|企業)[：:\s]+(.+)",
        "name": r"(?:名前|氏名|担当者)[：:\s]+(.+)",
        "collaboration": r"(?:協業|協業か)[：:\s]+(.+)",
        "prospect": r"(?:クライアント見込み|見込み|クライアント)[：:\s]+(.+)",
        "importance": r"(?:重要度|優先度)[：:\s]+(.+)",
        "memo": r"(?:メモ|備考|ノート)[：:\s]+(.+)",
    }
    for key, pattern in patterns.items():
        match = re.search(pattern, text)
        if match:
            fields[key] = match.group(1).strip()
    if "company" in fields or "name" in fields:
        return fields
    return None


def parse_delete(text):
    """削除コマンド解析"""
    if not re.search(r"削除|remove|delete", text, re.IGNORECASE):
        return None
    company_match = re.search(r"(?:会社名|会社)[：:\s]+(.+)", text)
    name_match = re.search(r"(?:名前|氏名)[：:\s]+(.+)", text)
    if company_match or name_match:
        return {
            "company": company_match.group(1).strip() if company_match else None,
            "name": name_match.group(1).strip() if name_match else None,
        }
    simple = re.search(r"(?:削除|remove|delete)\s+(.+)", text, re.IGNORECASE)
    if simple:
        keyword = simple.group(1).strip()
        return {"company": keyword, "name": keyword}
    return None


# ── Slackイベント処理 ──

@app.event("app_mention")
def handle_mention(event, say):
    """メンションに即座に応答"""
    text = event.get("text", "")
    thread_ts = event.get("thread_ts", event.get("ts"))

    # メンション部分を除去
    clean = re.sub(r"<@[A-Z0-9]+>", "", text).strip()

    # 追加
    add_data = parse_add(clean)
    if add_data:
        no = add_entry(
            add_data.get("company", ""),
            add_data.get("name", ""),
            add_data.get("collaboration", ""),
            add_data.get("prospect", ""),
            add_data.get("importance", "通常"),
            add_data.get("memo", ""),
        )
        say(
            text=(
                f"✅ ステークホルダーを登録しました！\n\n"
                f"*No.{no}*\n"
                f"• 会社名: {add_data.get('company', '')}\n"
                f"• 名前: {add_data.get('name', '')}\n"
                f"• 協業: {add_data.get('collaboration', '')}\n"
                f"• クライアント見込み: {add_data.get('prospect', '')}\n"
                f"• 重要度: {add_data.get('importance', '通常')}\n"
                f"• メモ: {add_data.get('memo', '')}"
            ),
            thread_ts=thread_ts,
        )
        return

    # 削除
    del_data = parse_delete(clean)
    if del_data:
        deleted = delete_entry(del_data.get("company"), del_data.get("name"))
        if deleted:
            items = "\n".join(f"• {d}" for d in deleted)
            say(text=f"🗑️ 以下を削除しました:\n{items}", thread_ts=thread_ts)
        else:
            keyword = del_data.get("company") or del_data.get("name") or "（不明）"
            say(text=f"⚠️ 「{keyword}」に該当するエントリが見つかりませんでした。", thread_ts=thread_ts)
        return

    # 一覧
    if re.search(r"一覧|リスト|list|表示|確認", clean, re.IGNORECASE):
        entries = get_all_entries()
        if not entries:
            say(text="📋 ステークホルダーリストは空です。", thread_ts=thread_ts)
            return
        lines = ["📋 *ステークホルダーリスト*\n"]
        for e in entries:
            imp = e["importance"]
            icon = "🔴" if imp in ["高", "最高"] else "🟡" if imp == "通常" else "⚪"
            lines.append(
                f"{icon} *No.{e['no']}* {e['company']} / {e['name']}\n"
                f"   協業: {e['collaboration']} | 見込み: {e['prospect']} | 重要度: {imp}\n"
                f"   メモ: {e['memo']}"
            )
        lines.append(f"\n合計: {len(entries)}件")
        say(text="\n".join(lines), thread_ts=thread_ts)
        return

    # ヘルプ
    say(
        text=(
            "📝 *営業リスト管理Bot* — 使い方\n\n"
            "*【追加】*\n```\n@営業リストBot 追加\n会社名: 株式会社ABC\n名前: 田中太郎\n"
            "協業: ○\nクライアント見込み: ◎\n重要度: 高\nメモ: 来月打ち合わせ予定```\n\n"
            "*【削除】*\n```\n@営業リストBot 削除 株式会社ABC```\n\n"
            "*【一覧表示】*\n```\n@営業リストBot 一覧```"
        ),
        thread_ts=thread_ts,
    )


def main():
    print("=== 営業リストBot (Socket Mode) 起動 ===")
    get_excel_path()  # 初期化
    handler = SocketModeHandler(app, SLACK_APP_TOKEN)
    handler.start()


if __name__ == "__main__":
    main()
