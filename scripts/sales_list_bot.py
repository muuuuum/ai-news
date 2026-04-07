#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Slack Bot: 辻倉_営業リスト報告チャンネルでステークホルダー管理
代表がSlackに投稿 → Excelに追加/削除
"""

import json
import sys
import os
import datetime
import urllib.request
import urllib.parse
import re

# openpyxl
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

TOKEN = sys.argv[1] if len(sys.argv) > 1 else os.environ.get("SLACK_BOT_TOKEN", "")
BOT_ID = "U0AQV9ADY6B"

# チャンネル設定（初回実行時にチャンネルIDを設定してください）
CHANNEL_NAME = "辻倉_営業リスト報告"
CHANNEL_ID = sys.argv[2] if len(sys.argv) > 2 else os.environ.get("SALES_CHANNEL_ID", "C0AR48QHRGB")

# Excelファイルパス
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(SCRIPT_DIR, "..", "stakeholder_list.xlsx")

# ヘッダー定義
HEADERS = ["No.", "会社名", "名前", "協業", "クライアント見込み", "重要度", "メモ", "登録日"]


def slack_api(method, data=None):
    """Slack API を呼び出す"""
    url = f"https://slack.com/api/{method}"
    headers = {
        "Authorization": f"Bearer {TOKEN}",
        "Content-Type": "application/json; charset=utf-8",
    }
    if data:
        body = json.dumps(data, ensure_ascii=False).encode("utf-8")
        req = urllib.request.Request(url, data=body, headers=headers, method="POST")
    else:
        req = urllib.request.Request(url, headers=headers)

    try:
        with urllib.request.urlopen(req) as resp:
            return json.loads(resp.read().decode("utf-8"))
    except Exception as e:
        print(f"API error ({method}): {e}")
        return {"ok": False, "error": str(e)}


def post_message(channel_id, text, thread_ts=None):
    """メッセージを投稿"""
    data = {"channel": channel_id, "text": text, "mrkdwn": True}
    if thread_ts:
        data["thread_ts"] = thread_ts
    result = slack_api("chat.postMessage", data)
    return result.get("ok", False)


def get_messages(channel_id, limit=20):
    """チャンネルのメッセージを取得"""
    result = slack_api(f"conversations.history?channel={channel_id}&limit={limit}")
    return result.get("messages", []) if result.get("ok") else []


def get_thread(channel_id, ts):
    """スレッド内のメッセージを取得"""
    result = slack_api(f"conversations.replies?channel={channel_id}&ts={ts}")
    return result.get("messages", []) if result.get("ok") else []


# ── Excel操作 ──

def init_excel():
    """Excelファイルを初期化（存在しない場合のみ）"""
    if os.path.exists(EXCEL_PATH):
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ステークホルダーリスト"

    # ヘッダースタイル
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 列幅設定
    widths = {"A": 6, "B": 25, "C": 15, "D": 10, "E": 18, "F": 10, "G": 40, "H": 14}
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    # フィルター設定
    ws.auto_filter.ref = f"A1:H1"

    wb.save(EXCEL_PATH)
    print(f"✅ Excelファイルを作成: {EXCEL_PATH}")


def load_workbook():
    """Excelファイルを読み込む"""
    init_excel()
    return openpyxl.load_workbook(EXCEL_PATH)


def get_next_no(ws):
    """次のNo.を取得"""
    max_no = 0
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0] and isinstance(row[0], (int, float)):
            max_no = max(max_no, int(row[0]))
    return max_no + 1


def add_entry(company, name, is_collaboration, is_prospect, importance, memo):
    """ステークホルダーをExcelに追加"""
    wb = load_workbook()
    ws = wb.active

    no = get_next_no(ws)
    jst = datetime.timezone(datetime.timedelta(hours=9))
    today = datetime.datetime.now(jst).strftime("%Y-%m-%d")

    new_row = [no, company, name, is_collaboration, is_prospect, importance, memo, today]
    row_num = ws.max_row + 1

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col, value in enumerate(new_row, 1):
        cell = ws.cell(row=row_num, column=col, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center")

    # フィルター範囲を更新
    ws.auto_filter.ref = f"A1:H{row_num}"

    wb.save(EXCEL_PATH)
    print(f"✅ 追加: No.{no} {company} - {name}")
    return no


def delete_entry(company=None, name=None):
    """ステークホルダーをExcelから削除（会社名 or 名前で検索）"""
    wb = load_workbook()
    ws = wb.active

    deleted = []
    rows_to_delete = []

    for row in range(2, ws.max_row + 1):
        row_company = ws.cell(row=row, column=2).value or ""
        row_name = ws.cell(row=row, column=3).value or ""

        match = False
        if company and name:
            match = company in str(row_company) and name in str(row_name)
        elif company:
            match = company in str(row_company)
        elif name:
            match = name in str(row_name)

        if match:
            deleted.append(f"No.{ws.cell(row=row, column=1).value} {row_company} - {row_name}")
            rows_to_delete.append(row)

    # 下の行から削除（行番号がずれないように）
    for row in reversed(rows_to_delete):
        ws.delete_rows(row)

    # No.を振り直す
    for i, row in enumerate(range(2, ws.max_row + 1), 1):
        ws.cell(row=row, column=1, value=i)

    if rows_to_delete:
        ws.auto_filter.ref = f"A1:H{ws.max_row}"

    wb.save(EXCEL_PATH)
    return deleted


def get_all_entries():
    """全エントリをリストで取得"""
    wb = load_workbook()
    ws = wb.active
    entries = []
    for row in range(2, ws.max_row + 1):
        entry = {
            "no": ws.cell(row=row, column=1).value,
            "company": ws.cell(row=row, column=2).value or "",
            "name": ws.cell(row=row, column=3).value or "",
            "collaboration": ws.cell(row=row, column=4).value or "",
            "prospect": ws.cell(row=row, column=5).value or "",
            "importance": ws.cell(row=row, column=6).value or "",
            "memo": ws.cell(row=row, column=7).value or "",
            "date": ws.cell(row=row, column=8).value or "",
        }
        entries.append(entry)
    return entries


# ── メッセージ解析 ──

def parse_add_command(text):
    """追加コマンドを解析

    フォーマット例:
    追加
    会社名: 株式会社ABC
    名前: 田中太郎
    協業: ○
    クライアント見込み: ◎
    重要度: 高
    メモ: 来月打ち合わせ予定
    """
    clean = text.replace(f"<@{BOT_ID}>", "").strip()

    if not re.search(r"追加|登録|add", clean, re.IGNORECASE):
        return None

    fields = {}
    patterns = {
        "company": r"(?:会社名|会社|企業)[：:\s]+(.+)",
        "name": r"(?:名前|氏名|担当者)[：:\s]+(.+)",
        "collaboration": r"(?:協業|協業か)[：:\s]+(.+)",
        "prospect": r"(?:クライアント見込み|見込み|クライアント)[：:\s]+(.+)",
        "importance": r"(?:重要度|優先度)[：:\s]+(.+)",
        "memo": r"(?:メモ|備考|ノート)[：:\s]+(.+)",
    }

    for key, pattern in patterns.items():
        match = re.search(pattern, clean)
        if match:
            fields[key] = match.group(1).strip()

    if "company" in fields or "name" in fields:
        return {
            "company": fields.get("company", ""),
            "name": fields.get("name", ""),
            "collaboration": fields.get("collaboration", ""),
            "prospect": fields.get("prospect", ""),
            "importance": fields.get("importance", "通常"),
            "memo": fields.get("memo", ""),
        }

    return None


def parse_delete_command(text):
    """削除コマンドを解析

    フォーマット例:
    削除 株式会社ABC
    削除 会社名: 株式会社ABC
    削除 名前: 田中太郎
    """
    clean = text.replace(f"<@{BOT_ID}>", "").strip()

    if not re.search(r"削除|remove|delete", clean, re.IGNORECASE):
        return None

    # パターン1: 「削除 会社名: XXX」
    company_match = re.search(r"(?:会社名|会社)[：:\s]+(.+)", clean)
    name_match = re.search(r"(?:名前|氏名)[：:\s]+(.+)", clean)

    if company_match or name_match:
        return {
            "company": company_match.group(1).strip() if company_match else None,
            "name": name_match.group(1).strip() if name_match else None,
        }

    # パターン2: 「削除 XXX」（会社名として扱う）
    simple_match = re.search(r"(?:削除|remove|delete)\s+(.+)", clean, re.IGNORECASE)
    if simple_match:
        keyword = simple_match.group(1).strip()
        return {"company": keyword, "name": keyword}

    return None


def parse_list_command(text):
    """一覧表示コマンドを判定"""
    clean = text.replace(f"<@{BOT_ID}>", "").strip()
    return bool(re.search(r"一覧|リスト|list|表示|確認", clean, re.IGNORECASE))


# ── メイン処理 ──

def find_unreplied_mentions(channel_id):
    """未返信のメンションを検索"""
    unreplied = []
    messages = get_messages(channel_id, 20)

    for msg in messages:
        user = msg.get("user", "")
        text = msg.get("text", "")
        ts = msg.get("ts", "")
        reply_count = msg.get("reply_count", 0)

        # トップレベルの未返信メンション
        if user != BOT_ID and f"<@{BOT_ID}>" in text and reply_count == 0:
            unreplied.append({"ts": ts, "parent_ts": ts, "text": text, "type": "top"})

        # スレッド内もチェック
        if reply_count > 0:
            thread = get_thread(channel_id, ts)
            for t_msg in thread:
                t_user = t_msg.get("user", "")
                t_text = t_msg.get("text", "")
                t_ts = t_msg.get("ts", "")

                if t_user != BOT_ID and f"<@{BOT_ID}>" in t_text:
                    t_ts_float = float(t_ts)
                    bot_replied = any(
                        float(x.get("ts", "0")) > t_ts_float and x.get("user") == BOT_ID
                        for x in thread
                    )
                    if not bot_replied:
                        unreplied.append({"ts": t_ts, "parent_ts": ts, "text": t_text, "type": "thread"})

    return unreplied


def process_mention(mention):
    """メンションを処理してExcel操作を実行"""
    text = mention["text"]

    # 追加コマンド
    add_data = parse_add_command(text)
    if add_data:
        no = add_entry(
            add_data["company"],
            add_data["name"],
            add_data["collaboration"],
            add_data["prospect"],
            add_data["importance"],
            add_data["memo"],
        )
        reply = (
            f"✅ ステークホルダーを登録しました！\n\n"
            f"*No.{no}*\n"
            f"• 会社名: {add_data['company']}\n"
            f"• 名前: {add_data['name']}\n"
            f"• 協業: {add_data['collaboration']}\n"
            f"• クライアント見込み: {add_data['prospect']}\n"
            f"• 重要度: {add_data['importance']}\n"
            f"• メモ: {add_data['memo']}\n"
        )
        return reply

    # 削除コマンド
    del_data = parse_delete_command(text)
    if del_data:
        deleted = delete_entry(del_data.get("company"), del_data.get("name"))
        if deleted:
            items = "\n".join(f"• {d}" for d in deleted)
            return f"🗑️ 以下を削除しました:\n{items}"
        else:
            keyword = del_data.get("company") or del_data.get("name") or "（不明）"
            return f"⚠️ 「{keyword}」に該当するエントリが見つかりませんでした。"

    # 一覧表示
    if parse_list_command(text):
        entries = get_all_entries()
        if not entries:
            return "📋 ステークホルダーリストは空です。"

        lines = ["📋 *ステークホルダーリスト*\n"]
        for e in entries:
            imp = e["importance"]
            imp_icon = "🔴" if imp in ["高", "最高"] else "🟡" if imp == "通常" else "⚪"
            lines.append(
                f"{imp_icon} *No.{e['no']}* {e['company']} / {e['name']}\n"
                f"   協業: {e['collaboration']} | 見込み: {e['prospect']} | 重要度: {imp}\n"
                f"   メモ: {e['memo']}"
            )
        lines.append(f"\n合計: {len(entries)}件")
        return "\n".join(lines)

    # ヘルプ
    return (
        "📝 *営業リスト管理Bot* — 使い方\n\n"
        "*【追加】*\n"
        "```\n"
        f"<@{BOT_ID}> 追加\n"
        "会社名: 株式会社ABC\n"
        "名前: 田中太郎\n"
        "協業: ○\n"
        "クライアント見込み: ◎\n"
        "重要度: 高\n"
        "メモ: 来月打ち合わせ予定\n"
        "```\n\n"
        "*【削除】*\n"
        f"```\n<@{BOT_ID}> 削除 株式会社ABC```\n\n"
        "*【一覧表示】*\n"
        f"```\n<@{BOT_ID}> 一覧```"
    )


def main():
    """メイン処理"""
    jst = datetime.timezone(datetime.timedelta(hours=9))
    now = datetime.datetime.now(jst).isoformat()
    print(f"=== 営業リストBot started at {now} ===")

    if not CHANNEL_ID:
        print("❌ CHANNEL_ID が設定されていません。")
        print("   引数または環境変数 SALES_CHANNEL_ID で指定してください。")
        sys.exit(1)

    # Excel初期化
    init_excel()

    # メンション処理
    unreplied = find_unreplied_mentions(CHANNEL_ID)
    print(f"未返信メンション: {len(unreplied)}件")

    for mention in unreplied:
        reply_text = process_mention(mention)
        ok = post_message(CHANNEL_ID, reply_text, mention["parent_ts"])
        status = "✅" if ok else "❌"
        print(f"{status} replied: {mention['text'][:60]}")

    print("=== Done ===")


if __name__ == "__main__":
    main()
